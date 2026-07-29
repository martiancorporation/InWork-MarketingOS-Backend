"""Unit tests for the CSV/Excel/PDF report renderers against a small fixed
``ReportContent`` fixture. ``render_visual`` needs a real headless browser and
is exercised instead (faked) in ``tests/integration/test_reports.py``.
"""

from __future__ import annotations

import csv
import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.schemas.analytics import AnalyticsTotals
from app.services.reports.content import CampaignRow, ChannelRow, ReportContent
from app.services.reports.render_csv import render_csv
from app.services.reports.render_excel import render_excel
from app.services.reports.render_pdf import render_pdf


@pytest.fixture
def content() -> ReportContent:
    totals = AnalyticsTotals(
        impressions=1000,
        clicks=50,
        conversions=5,
        leads=5,
        spend=100.0,
        revenue=200.0,
        ctr=5.0,
        cpl=20.0,
        roas=2.0,
    )
    return ReportContent(
        client_name="Acme <Co> & Sons",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 29),
        channel_labels=["Meta", "Google Ads"],
        totals=totals,
        channel_breakdown=[ChannelRow(label="Meta", totals=totals)],
        campaigns=[
            CampaignRow(
                name="Summer <Sale>",
                status="active",
                spend=100.0,
                leads=5,
                cpl=20.0,
                target_cpl=25.0,
                ctr=5.0,
                target_ctr=4.0,
            )
        ],
        top_campaigns=[
            CampaignRow(
                name="Summer <Sale>",
                status="active",
                spend=100.0,
                leads=5,
                cpl=20.0,
                target_cpl=25.0,
                ctr=5.0,
                target_ctr=4.0,
            )
        ],
        went_right="Summer <Sale> led with 5 leads.",
        went_wrong="Nothing went wrong.",
    )


def test_render_csv_parses_back(content: ReportContent) -> None:
    out = render_csv(content)
    assert out
    text = out.decode("utf-8")
    rows = list(csv.reader(io.StringIO(text)))
    assert any("### Campaign Performance" in r for r in rows)
    assert any("Summer <Sale>" in r for r in rows)


def test_render_csv_handles_empty_content() -> None:
    empty = ReportContent(
        client_name="Quiet Co.",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 29),
        channel_labels=[],
        totals=AnalyticsTotals(),
    )
    out = render_csv(empty)
    assert b"No campaigns in this period." in out
    assert b"No channel data in this period." in out


def test_render_excel_loads_back(content: ReportContent) -> None:
    out = render_excel(content)
    assert out
    wb = load_workbook(io.BytesIO(out))
    assert "Summary" in wb.sheetnames
    assert "Campaign Performance" in wb.sheetnames
    ws = wb["Campaign Performance"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "Summer <Sale>" in values


def test_render_excel_handles_empty_content() -> None:
    empty = ReportContent(
        client_name="Quiet Co.",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 29),
        channel_labels=[],
        totals=AnalyticsTotals(),
    )
    out = render_excel(empty)
    wb = load_workbook(io.BytesIO(out))
    assert "Campaign Performance" in wb.sheetnames


def test_render_pdf_produces_a_valid_pdf(content: ReportContent) -> None:
    out = render_pdf(content)
    assert out[:5] == b"%PDF-"
    assert len(out) > 500


def test_render_pdf_escapes_html_special_characters(content: ReportContent) -> None:
    # The client name contains "<Co> & Sons" — must not crash reportlab's
    # Paragraph markup parser (Table cells are plain strings and don't need
    # escaping; the client name goes through a Paragraph in the title).
    out = render_pdf(content)
    assert out[:5] == b"%PDF-"


def test_render_pdf_handles_empty_content() -> None:
    empty = ReportContent(
        client_name="Quiet Co.",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 29),
        channel_labels=[],
        totals=AnalyticsTotals(),
    )
    out = render_pdf(empty)
    assert out[:5] == b"%PDF-"
