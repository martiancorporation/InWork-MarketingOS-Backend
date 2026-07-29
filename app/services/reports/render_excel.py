"""Render a :class:`ReportContent` as an .xlsx workbook — one worksheet per
requested section, using ``openpyxl`` (already a pinned dependency).

``openpyxl`` is imported lazily, inside the functions that need it, not at
module top level: it pulls in ``numpy`` transitively, and importing this
module is on the app's startup import path (via ``generator.py`` →
``report_service.py`` → the reports router). A numpy build incompatible with
the host CPU must only break Excel export, not boot the whole app — mirrors
how ``boto3``/``playwright`` are already lazily imported elsewhere in this repo.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

from app.services.reports.content import ReportContent

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

_SECTION_TITLES = {
    "campaign_performance": "Campaign Performance",
    "ga_overview": "Channel Overview",
    "top_ads": "Top Campaigns",
    "went_wrong_right": "Summary",
}

_CAMPAIGN_HEADER = [
    "Campaign",
    "Status",
    "Spend",
    "Leads",
    "CPL",
    "Target CPL",
    "CTR %",
    "Target CTR %",
]
_CHANNEL_HEADER = [
    "Channel",
    "Impressions",
    "Clicks",
    "Conversions",
    "Leads",
    "Spend",
    "Revenue",
    "CTR %",
    "CPL",
    "ROAS",
]


def _write_header(ws: Worksheet, header: list[str]) -> None:
    from openpyxl.styles import Font

    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet) -> None:
    """openpyxl doesn't size columns to content — do it ourselves, cheaply."""
    from openpyxl.utils import get_column_letter

    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 40)


def _campaign_sheet(wb: Workbook, title: str, rows) -> None:
    ws = wb.create_sheet(title)
    _write_header(ws, _CAMPAIGN_HEADER)
    if not rows:
        ws.append(["No campaigns in this period."])
    for c in rows:
        ws.append([c.name, c.status, c.spend, c.leads, c.cpl, c.target_cpl, c.ctr, c.target_ctr])
    _autosize(ws)


def render_excel(content: ReportContent) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    summary = wb.create_sheet("Summary")
    summary.append([f"{content.client_name} — Report"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([f"{content.date_from.isoformat()} to {content.date_to.isoformat()}"])
    summary.append([f"Channels: {', '.join(content.channel_labels) or '(none)'}"])
    if "went_wrong_right" in content.included_sections:
        summary.append([])
        summary.append(["What went right", content.went_right])
        summary.append(["What went wrong", content.went_wrong])
    _autosize(summary)

    if "campaign_performance" in content.included_sections:
        _campaign_sheet(wb, _SECTION_TITLES["campaign_performance"], content.campaigns)

    if "top_ads" in content.included_sections:
        _campaign_sheet(wb, _SECTION_TITLES["top_ads"], content.top_campaigns)

    if "ga_overview" in content.included_sections:
        ws = wb.create_sheet(_SECTION_TITLES["ga_overview"])
        _write_header(ws, _CHANNEL_HEADER)
        if not content.channel_breakdown:
            ws.append(["No channel data in this period."])
        for row in content.channel_breakdown:
            t = row.totals
            ws.append(
                [
                    row.label,
                    t.impressions,
                    t.clicks,
                    t.conversions,
                    t.leads,
                    t.spend,
                    t.revenue,
                    t.ctr,
                    t.cpl,
                    t.roas,
                ]
            )
        t = content.totals
        ws.append(
            [
                "Total",
                t.impressions,
                t.clicks,
                t.conversions,
                t.leads,
                t.spend,
                t.revenue,
                t.ctr,
                t.cpl,
                t.roas,
            ]
        )
        _autosize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
